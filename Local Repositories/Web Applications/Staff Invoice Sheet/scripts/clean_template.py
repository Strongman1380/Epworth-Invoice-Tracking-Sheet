#!/usr/bin/env python3
"""
Clean the billing template: clear data rows (9-96) of input-only columns,
clear header personal info, but preserve all formulas, formatting, validations,
named ranges, and table objects.
"""

import os
import shutil
from openpyxl import load_workbook

SRC = os.path.join(os.path.dirname(__file__), '..', 'Dean.Raquel.TS.012426.xlsx')
DST = os.path.join(os.path.dirname(__file__), '..', 'templates', 'timesheet-template.xlsx')

# Input-only columns to clear (1-indexed): A-D, F-L, M-P, R-V
# Formula columns to KEEP: E(5), Q(17), W(23), X(24), Y(25), Z(26), AA(27), AB(28), AC(29)
INPUT_COLS = [1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22]

# Header cells to clear (employee-specific info)
HEADER_CELLS_TO_CLEAR = [
    'Q2',   # Employee name
    'Q4',   # Supervisor name
    'Q6',   # Pay period
    'S5',   # Employee signature
    'W6',   # Supervisor signature (if filled)
]

wb = load_workbook(SRC)
ws = wb['Sheet1']

# Clear header personal info
for cell_ref in HEADER_CELLS_TO_CLEAR:
    cell = ws[cell_ref]
    cell.value = None

# Clear data rows (9-96) - only input columns
for row in range(9, 97):
    for col in INPUT_COLS:
        cell = ws.cell(row=row, column=col)
        # Only clear if NOT a formula
        val = cell.value
        if val is not None and not (isinstance(val, str) and val.startswith('=')):
            cell.value = None

# Also clear totals row input cells (row 97) - leave formula cells
# Row 97 columns with SUM formulas should stay
# Clear only the non-formula cells in row 97
for col in INPUT_COLS:
    cell = ws.cell(row=97, column=col)
    val = cell.value
    if val is not None and not (isinstance(val, str) and val.startswith('=')):
        cell.value = None

wb.save(DST)
print(f"SUCCESS: Created clean template at {DST}")
print(f"File size: {os.path.getsize(DST)} bytes")

# Verify formulas are preserved
wb2 = load_workbook(DST)
ws2 = wb2['Sheet1']
print(f"\nVerification - Formula in Q9: {ws2['Q9'].value}")
print(f"Verification - Formula in W9: {ws2['W9'].value}")
print(f"Verification - Formula in X9: {ws2['X9'].value}")
print(f"Verification - Formula in AC9: {ws2['AC9'].value}")
print(f"Verification - Formula in K97: {ws2['K97'].value}")
print(f"Verification - Q2 (name cleared): '{ws2['Q2'].value}'")
print(f"Verification - A9 (data cleared): '{ws2['A9'].value}'")
print(f"\nSheet names: {wb2.sheetnames}")
print(f"Data validations on Sheet1: {len(ws2.data_validations.dataValidation)}")
