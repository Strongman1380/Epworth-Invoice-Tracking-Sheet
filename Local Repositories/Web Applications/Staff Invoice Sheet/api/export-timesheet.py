"""
Vercel Python Serverless Function: Export timesheet data into the billing Excel template.
Loads the template with openpyxl, fills in data rows (preserving formulas), and returns
the completed .xlsx as a binary download.
"""

import json
import os
import io
from datetime import datetime, time
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'timesheet-template.xlsx')

# Data rows in the template: rows 9 through 96 (88 rows max)
DATA_ROW_START = 9
DATA_ROW_END = 96

# Formula columns that must NOT be overwritten (1-indexed)
# E=5 (VLOOKUP), Q=17 (Transport Bill Time), W=23 through AC=29
FORMULA_COLS = {5, 17, 23, 24, 25, 26, 27, 28, 29}

# Column mapping: web field name -> Excel column index (1-indexed)
COL_MAP = {
    'serviceType':    1,   # A
    'serviceChoice':  2,   # B
    'travelComments': 3,   # C
    'clientName':     4,   # D
    'caseNumber':     5,   # E - overrides VLOOKUP with hard value
    'dateOfService':  6,   # F
    'timeIn':         7,   # G
    'timeOut':        8,   # H
    'driveFrom':      9,   # I
    'driveTo':        10,  # J
    'miles':          11,  # K
    'staffName':      12,  # L
    'paidLeave':      13,  # M
    'paidHours':      14,  # N
    'meetingHours':   15,  # O
    'trainingHours':  16,  # P
    # Q=17 is Transport Bill Time formula - SKIP
    'driveNonBill':   18,  # R
    'docTime':        19,  # S
    'mileageAdj':     20,  # T
    'odometerStart':  21,  # U
    'odometerStop':   22,  # V
    # W-AC = formulas - SKIP
}

# Numeric fields that should be converted to float
NUMERIC_FIELDS = {
    'miles', 'paidHours', 'meetingHours', 'trainingHours',
    'driveNonBill', 'docTime', 'mileageAdj', 'odometerStart', 'odometerStop'
}


def parse_date(val):
    """Convert date string to datetime object for Excel."""
    if not val:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%y', '%m/%d/%Y'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return val  # Return as string if unparseable


def parse_time(val):
    """Convert time string (HH:MM) to a time object for Excel."""
    if not val:
        return None
    try:
        parts = val.split(':')
        return time(int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        return val


def parse_numeric(val):
    """Convert to float, return None if empty/invalid."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def fill_workbook(wb, payload):
    """Fill the template workbook with data from the payload."""
    ws = wb['Sheet1']

    # Fill header info
    header = payload.get('header', {})
    if header.get('employeeName'):
        ws['Q2'] = header['employeeName']
    if header.get('supervisorName'):
        ws['Q4'] = header['supervisorName']
    if header.get('payPeriod'):
        ws['Q6'] = header['payPeriod']
    if header.get('employeeSignature'):
        ws['S5'] = header['employeeSignature']
    if header.get('supervisorSignature'):
        ws['W6'] = header['supervisorSignature']

    # Fill data rows
    rows = payload.get('rows', [])
    for i, row_data in enumerate(rows):
        if i >= (DATA_ROW_END - DATA_ROW_START + 1):
            break  # Max 88 rows

        excel_row = DATA_ROW_START + i

        for field, col in COL_MAP.items():
            val = row_data.get(field)
            if val is None or val == '':
                continue

            # For caseNumber (col E), we hard-write the value
            # which overrides the VLOOKUP formula
            if col == 5:
                num_val = parse_numeric(val)
                if num_val is not None:
                    ws.cell(row=excel_row, column=col, value=int(num_val))
                else:
                    ws.cell(row=excel_row, column=col, value=val)
                continue

            # Skip formula columns (except E which we handle above)
            if col in FORMULA_COLS:
                continue

            # Date field
            if field == 'dateOfService':
                parsed = parse_date(val)
                if parsed:
                    cell = ws.cell(row=excel_row, column=col, value=parsed)
                    cell.number_format = 'MM/DD/YY'
                continue

            # Time fields
            if field in ('timeIn', 'timeOut'):
                parsed = parse_time(val)
                if parsed:
                    cell = ws.cell(row=excel_row, column=col, value=parsed)
                    cell.number_format = 'h:mm'
                continue

            # Numeric fields
            if field in NUMERIC_FIELDS:
                num_val = parse_numeric(val)
                if num_val is not None:
                    ws.cell(row=excel_row, column=col, value=num_val)
                continue

            # String fields
            ws.cell(row=excel_row, column=col, value=val)

    return wb


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            payload = json.loads(body)

            # Load template
            wb = load_workbook(TEMPLATE_PATH)

            # Fill data
            fill_workbook(wb, payload)

            # Write to bytes buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_bytes = output.getvalue()

            # Build filename
            employee = payload.get('header', {}).get('employeeName', 'Unnamed')
            pay_period = payload.get('header', {}).get('payPeriod', 'No_Date')
            # Clean filename
            safe_name = ''.join(c if c.isalnum() or c in (' ', '-', '_', '.') else '_' for c in employee)
            safe_period = ''.join(c if c.isalnum() or c in (' ', '-', '_', '.') else '_' for c in pay_period)
            filename = f"{safe_name}_{safe_period}_Timesheet.xlsx"

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', str(len(file_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(file_bytes)

        except json.JSONDecodeError:
            self.send_response(400)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Invalid JSON payload'}).encode())

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())
